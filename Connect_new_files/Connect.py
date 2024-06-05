from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font

# List of your Excel files
file_list = ["Task2_Done.xlsx", "Tables_Tasks_1345.xlsx"]

# Load the existing workbook where you want to merge the data
existing_file = "All_Tables_June5.xlsx"
book = load_workbook(existing_file)

def copy_sheet(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = Font(name=cell.font.name,
                                     size=cell.font.size,
                                     bold=cell.font.bold,
                                     italic=cell.font.italic,
                                     vertAlign=cell.font.vertAlign,
                                     underline=cell.font.underline,
                                     strike=cell.font.strike,
                                     color=cell.font.color)
                new_cell.border = Border(left=cell.border.left,
                                         right=cell.border.right,
                                         top=cell.border.top,
                                         bottom=cell.border.bottom,
                                         diagonal=cell.border.diagonal,
                                         diagonal_direction=cell.border.diagonal_direction,
                                         outline=cell.border.outline,
                                         vertical=cell.border.vertical,
                                         horizontal=cell.border.horizontal)
                new_cell.fill = PatternFill(fill_type=cell.fill.fill_type,
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color)
                new_cell.number_format = cell.number_format
                new_cell.protection = Protection(locked=cell.protection.locked,
                                                 hidden=cell.protection.hidden)
                new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                               vertical=cell.alignment.vertical,
                                               text_rotation=cell.alignment.text_rotation,
                                               wrap_text=cell.alignment.wrap_text,
                                               shrink_to_fit=cell.alignment.shrink_to_fit,
                                               indent=cell.alignment.indent)

    for col in source_sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        target_sheet.column_dimensions[column].width = adjusted_width

    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

for file in file_list:
    wb = load_workbook(file)
    for sheet_name in wb.sheetnames:
        source_sheet = wb[sheet_name]
        new_sheet = book.create_sheet(sheet_name)
        copy_sheet(source_sheet, new_sheet)

book.save(existing_file)
